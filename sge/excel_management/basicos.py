from statistics import median
from tkinter.scrolledtext import example

import openpyxl as ox


# woorbook -> conjunto de hojas
# sheet -> hoja de calculo
# columnas (A,B,C...) y filas (1,2,3...)
# cell -> cuadrado de la tabla
"""
woorkook = ox.load_workbook("ejemplo.xlsx")
print(type(woorkook))
hojas = woorkook.sheetnames
notas_sheet = woorkook["Notas"]
print(notas_sheet)
#hojaEnUso = woorbook.activate -> Te da la hoja on top, en activo

alumno = notas_sheet["A2"]
print(notas_sheet["A2"].value)

if alumno.value == "Jesus":
    print("Yes")
    print(alumno.coordinate)

custom_sheet = woorkook["NombreCustom"]
cel = custom_sheet.cell(1,27)
print(cel.value)
"""

woorbook = ox.load_workbook("ejemplo.xlsx")
print(type(woorbook))
hojas = woorbook.sheetnames
nota_sheet = woorbook["Notas"]
print(nota_sheet)

print(f"Filas: {nota_sheet.min_row}")
print(f"Filas: {nota_sheet.max_row}")
print(f"Columnas: {nota_sheet.min_column}")
print(f"Columnas: {nota_sheet.max_column}")

# Calcular y motrar las medias de los alumnos con su nombre
# Los ejercicios cuentan un 50% y el examen otro 50%

for fila in range (2, nota_sheet.max_row + 1):
    ej1 = nota_sheet.cell(row=fila, column=2).value
    ej2 = nota_sheet.cell(row=fila, column=3).value
    ej3 = nota_sheet.cell(row=fila, column=4).value
    examen = nota_sheet.cell(row=fila, column=5).value

    media = (ej1, ej2, ej3)*0.5 / 3 + examen * 0.5
    print(f"media{media}")
