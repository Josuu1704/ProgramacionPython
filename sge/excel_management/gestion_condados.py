import openpyxl

#condados_woorbook = openpyxl.load_workbook("censuspopdata.xlsx")
#
#hoja_condados = condados_woorbook.active
#print(hoja_condados)
#
#print(f"Columnas inicio: {hoja_condados.min_column}")
#print(f"Columnas fin: {hoja_condados.max_column}")
#print(f"Filas inicio: {hoja_condados.min_row}")
#print(f"Filas fin: {hoja_condados.max_row}")
#
#datos ={
#  'Texas': {
#      'condados': {
#          'Travis': {
#              'sectores': 1,
#              'poblacion': 100
#          }
#      },
#      'poblacion_total': 10000
#  }
#}

#var = {
#    "Pais": "Estados Unidos de America",
#    "Poblacion total": 2,
#    "Estados": [
#        {
#            "nombre": "WY",
#            "poblacionTotalEstado": 4,
#            "Condados": [
#                {
#                    "Lincoln": {
#                        "Poblacion_Condado": 3,
#                        "Poblacion_Condado": 2,
#                    }
#                },
#                {
#                    "Matrona": {
#                        "Poblacion_Condado": 3,
#                        "Secciones:Condado": 2,
#                    }
#                }
#            ]
#        },
#
#    ]
#}

import openpyxl
import pprint

condados_woorbook = openpyxl.load_workbook("censuspopdata.xlsx")
hoja_condados = condados_woorbook.active

datos = {
    "Pais": "Estados Unidos de America",
    "Poblacion total": 0,
    "Estados": {}
}

for fila in range(hoja_condados.min_row + 1, hoja_condados.max_row + 1):

    estado = hoja_condados.cell(row=fila, column=2).value
    condado = hoja_condados.cell(row=fila, column=3).value
    poblacion = hoja_condados.cell(row=fila, column=4).value

    if estado not in datos["Estados"]:
        datos["Estados"][estado] = {
            "nombre": estado,
            "poblacionTotalEstado": 0,
            "Condados": {}
        }

    if condado not in datos["Estados"][estado]["Condados"]:
        datos["Estados"][estado]["Condados"][condado] = {
            "Poblacion_Condado": 0,
            "Secciones_Condado": 0
        }

    datos["Estados"][estado]["Condados"][condado]["Poblacion_Condado"] += poblacion
    datos["Estados"][estado]["Condados"][condado]["Secciones_Condado"] += 1
    datos["Estados"][estado]["poblacionTotalEstado"] += poblacion
    datos["Poblacion total"] += poblacion

estructura_final = {
    "Pais": datos["Pais"],
    "Poblacion total": datos["Poblacion total"],
    "Estados": []
}

for estado in datos["Estados"].values():
    lista_condados = []

    for nombre_condado, info_condado in estado["Condados"].items():
        lista_condados.append({
            nombre_condado: info_condado
        })

    estructura_final["Estados"].append({
        "nombre": estado["nombre"],
        "poblacionTotalEstado": estado["poblacionTotalEstado"],
        "Condados": lista_condados
    })
pprint.pprint(estructura_final)

with open("estructura_censo.txt", "w", encoding="utf-8") as archivo:
    archivo.write(pprint.pformat(estructura_final))
